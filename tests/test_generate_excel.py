"""generate_excel.py 的单元测试。"""
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent / "scripts"))

from generate_excel import generate


@pytest.fixture
def sample_results():
    """最小可用的分类结果样本（v1.1：双轴分类——type + advice）。"""
    return [
        {
            "keyword": "zerogpt",
            "translation": "零 GPT",
            "intent": "导航型",
            "type": "工具",
            "monetization": "SaaS",
            "advice": "待研究",
            "target_user": "学生/内容创作者/编辑",
            "urls": {
                "intitle": "https://www.google.com/search?q=intitle:%22zerogpt%22&hl=en&gl=us",
                "serp": "https://www.google.com/search?q=zerogpt&hl=en&gl=us",
                "trends": "https://trends.google.com/trends/explore?q=zerogpt&geo=US&date=today%201-m",
                "ahrefs": "https://ahrefs.com/keyword-difficulty/?country=us&input=zerogpt",
            },
        },
        {
            "keyword": "buy nike shoes",
            "translation": "[品牌库]",
            "intent": "交易型",
            "type": "带货",
            "monetization": "联盟",
            "advice": "跳过",
            "target_user": "[品牌库:nike] 鞋类消费者",
            "urls": {
                "intitle": "https://www.google.com/search?q=intitle:%22buy%20nike%20shoes%22&hl=en&gl=us",
                "serp": "https://www.google.com/search?q=buy%20nike%20shoes&hl=en&gl=us",
                "trends": "https://trends.google.com/trends/explore?q=buy%20nike%20shoes&geo=US&date=today%201-m",
                "ahrefs": "https://ahrefs.com/keyword-difficulty/?country=us&input=buy%20nike%20shoes",
            },
        },
        {
            "keyword": "thunderstorm tracker uk",
            "translation": "英国雷暴追踪器",
            "intent": "工具型",
            "type": "工具",
            "monetization": "SaaS / AdSense",
            "advice": "建议",
            "target_user": "英国户外活动者/气象爱好者",
            "urls": {
                "intitle": "https://www.google.com/search?q=intitle:%22thunderstorm%20tracker%20uk%22&hl=en&gl=us",
                "serp": "https://www.google.com/search?q=thunderstorm%20tracker%20uk&hl=en&gl=us",
                "trends": "https://trends.google.com/trends/explore?q=thunderstorm%20tracker%20uk&geo=US&date=today%201-m",
                "ahrefs": "https://ahrefs.com/keyword-difficulty/?country=us&input=thunderstorm%20tracker%20uk",
            },
        },
        {
            "keyword": "jmail.world",
            "translation": "Jmail 世界",
            "intent": "信息型",
            "type": "空",
            "monetization": "待定",
            "advice": "待研究",
            "target_user": "爱泼斯坦案关注者/调查新闻读者",
            "urls": {
                "intitle": "https://www.google.com/search?q=intitle:%22jmail.world%22&hl=en&gl=us",
                "serp": "https://www.google.com/search?q=jmail.world&hl=en&gl=us",
                "trends": "https://trends.google.com/trends/explore?q=jmail.world&geo=US&date=today%201-m",
                "ahrefs": "https://ahrefs.com/keyword-difficulty/?country=us&input=jmail.world",
            },
        },
    ]


class TestGenerate:
    def test_创建文件(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        result_path = generate(sample_results, out)
        assert result_path == out
        assert out.exists()

    def test_13列(self, sample_results, tmp_path: Path):
        """v1.4：表格列数从 12 增至 13。"""
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        assert ws.max_column == 13

    def test_表头内容(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        headers = [ws.cell(row=1, column=i).value for i in range(1, 14)]
        assert headers == [
            "英文原文", "中文翻译", "搜索意图", "类型", "变现路径",
            "建议", "目标用户", "收藏", "我的调研结论",   # v1.4 新增"我的调研结论"
            "intitle 检查", "SERP 检查", "谷歌趋势", "Ahrefs KD",
        ]

    def test_数据行数(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        # 1 行表头 + 4 行数据
        assert ws.max_row == 5

    def test_数据内容(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        # 第 2 行：zerogpt → advice=待研究
        assert ws.cell(row=2, column=1).value == "zerogpt"
        assert ws.cell(row=2, column=6).value == "待研究"
        # 第 4 行：thunderstorm → type=工具
        assert ws.cell(row=4, column=4).value == "工具"

    def test_类型空_单元格留空(self, sample_results, tmp_path: Path):
        """类型哨兵值 "空" 在 Excel 中应为空字符串，不写字面字符。"""
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        # 第 5 行：jmail.world → type="空"
        cell = ws.cell(row=5, column=4)
        assert cell.value in (None, ""), f"应为空白，实际为 {cell.value!r}"
        # 其他列不受影响
        assert ws.cell(row=5, column=1).value == "jmail.world"
        assert ws.cell(row=5, column=6).value == "待研究"

    def test_冻结面板(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        # C2 表示冻结首行 + 第 1、2 列
        assert ws.freeze_panes == "C2"

    def test_筛选器(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        # v1.4：列数 13，筛选范围 A1:M1
        assert ws.auto_filter.ref is not None
        assert ws.auto_filter.ref == "A1:M1"

    def test_超链接(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        # 第 2 行第 10 列（intitle，v1.4 后右移到第 10 列）应有 hyperlink
        cell = ws.cell(row=2, column=10)
        assert cell.hyperlink is not None
        assert cell.hyperlink.target == sample_results[0]["urls"]["intitle"]

    def test_自动创建目录(self, sample_results, tmp_path: Path):
        out = tmp_path / "nested" / "deep" / "out.xlsx"
        generate(sample_results, out)
        assert out.exists()

    def test_空结果_只写表头(self, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate([], out)
        wb = load_workbook(out)
        ws = wb.active
        assert ws.max_row == 1
        assert ws.max_column == 13

    def test_Sheet名(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        wb = load_workbook(out)
        assert wb.active.title == "关键词分类"


class TestRowColoring:
    """验证行级着色（按「建议」）。v1.4 起覆盖范围扩展到 1~9 列（含调研列，不含收藏列）。"""

    def test_建议行_第1列浅绿(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        # 第 4 行：thunderstorm tracker uk → 建议
        cell = ws.cell(row=4, column=1)
        assert cell.fill.fgColor.rgb in ("00E2EFDA", "E2EFDA", "FFE2EFDA")

    def test_建议行_调研列也浅绿(self, sample_results, tmp_path: Path):
        """v1.4：行级着色覆盖到第 9 列（我的调研结论）。"""
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        cell = ws.cell(row=4, column=9)  # 第 9 列 = 我的调研结论
        assert cell.fill.fgColor.rgb in ("00E2EFDA", "E2EFDA", "FFE2EFDA")

    def test_待研究行_第1列浅黄(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        # 第 2 行：zerogpt → 待研究
        cell = ws.cell(row=2, column=1)
        assert cell.fill.fgColor.rgb in ("00FFF2CC", "FFF2CC", "FFFFF2CC")

    def test_待研究行_调研列也浅黄(self, sample_results, tmp_path: Path):
        """v1.4：行级着色覆盖到第 9 列。"""
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        cell = ws.cell(row=2, column=9)
        assert cell.fill.fgColor.rgb in ("00FFF2CC", "FFF2CC", "FFFFF2CC")

    def test_跳过行_第1列浅灰(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        # 第 3 行：buy nike shoes → 跳过
        cell = ws.cell(row=3, column=1)
        assert cell.fill.fgColor.rgb in ("00F2F2F2", "F2F2F2", "FFF2F2F2")

    def test_跳过行_调研列也浅灰(self, sample_results, tmp_path: Path):
        """v1.4：行级着色覆盖到第 9 列。"""
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        cell = ws.cell(row=3, column=9)
        assert cell.fill.fgColor.rgb in ("00F2F2F2", "F2F2F2", "FFF2F2F2")


class TestLinkColumnColoring:
    """验证链接列着色（覆盖行级）。v1.4 后链接列在 10~13 列。"""

    def test_intitle列_浅蓝(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        cell = ws.cell(row=2, column=10)
        assert cell.fill.fgColor.rgb in ("00DDEBF7", "DDEBF7", "FFDDEBF7")

    def test_SERP列_浅紫(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        cell = ws.cell(row=2, column=11)
        assert cell.fill.fgColor.rgb in ("00E4D7F0", "E4D7F0", "FFE4D7F0")

    def test_趋势列_浅橙(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        cell = ws.cell(row=2, column=12)
        assert cell.fill.fgColor.rgb in ("00FCE4D6", "FCE4D6", "FFFCE4D6")

    def test_Ahrefs列_浅绿(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        cell = ws.cell(row=2, column=13)
        assert cell.fill.fgColor.rgb in ("00E2EFDA", "E2EFDA", "FFE2EFDA")


class TestFavoriteColumn:
    """验证"收藏"列的固定深绿背景 + 白字加粗 + 列宽。v1.4 后仍位于第 8 列（H 列），不右移。"""

    def test_收藏列有数据验证(self, sample_results, tmp_path: Path):
        """v1.5：下拉框含'收藏'和'丢弃'两个选项。"""
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        dvs = list(ws.data_validations.dataValidation)
        assert len(dvs) >= 1
        dv = dvs[0]
        assert dv.type == "list"
        # v1.5：应同时含"收藏"和"丢弃"
        assert "收藏" in dv.formula1
        assert "丢弃" in dv.formula1
        # v1.4：收藏列仍在 H 列，DataValidation 范围 H2:Hx
        ranges = " ".join(str(s) for s in dv.sqref.ranges) if dv.sqref else ""
        assert "H2" in ranges, f"DataValidation 范围应包含 H2，实际 {ranges!r}"

    def test_收藏列默认空(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        # 所有数据行的"收藏"列（第 8 列 H 列）默认为空
        for row_idx in range(2, ws.max_row + 1):
            assert ws.cell(row=row_idx, column=8).value in (None, "")

    def test_收藏列固定宽度8(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        # v1.4：收藏列仍在第 8 列（H 列）
        assert ws.column_dimensions["H"].width == 8

    def test_收藏列_固定深绿背景_548235(self, sample_results, tmp_path: Path):
        """v1.3：第 8 列（H 列）所有数据行都应有深绿 #548235 实色填充。"""
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=8)
            assert cell.fill.patternType == "solid", \
                f"行{row_idx}: 应为实色填充，实际 {cell.fill.patternType}"
            fg = str(cell.fill.fgColor.rgb).upper()
            assert "548235" in fg, f"行{row_idx}: 应为 548235，实际 {fg}"

    def test_收藏列_白字加粗(self, sample_results, tmp_path: Path):
        """深绿背景下文字必须白色加粗，否则不可读。"""
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        cell = ws.cell(row=2, column=8)
        assert cell.font.color is not None
        color_rgb = str(cell.font.color.rgb).upper() if cell.font.color.rgb else ""
        assert "FFFFFF" in color_rgb or "FFFFFFFF" in color_rgb, \
            f"应为白色，实际 {color_rgb}"
        assert cell.font.b is True, "应加粗"

    def test_收藏列_无条件格式(self, sample_results, tmp_path: Path):
        """v1.3：移除条件格式，避免 toggle 颜色不稳 bug。"""
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        wb = load_workbook(out)
        # 整个工作簿应没有任何条件格式规则
        cf_count = sum(
            len(rules)
            for ws in wb.worksheets
            for rules in ws.conditional_formatting._cf_rules.values()
        )
        assert cf_count == 0, f"应无条件格式，实际 {cf_count} 条"

    def test_链接列固定宽度12(self, sample_results, tmp_path: Path):
        """v1.4：链接列位于 J/K/L/M 列。"""
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        for col_letter in ["J", "K", "L", "M"]:
            assert ws.column_dimensions[col_letter].width == 12

    def test_英文原文列固定30(self, sample_results, tmp_path: Path):
        """英文原文列与中文翻译列同宽，避免长英文词撑爆列宽。"""
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        assert ws.column_dimensions["A"].width == 30


class TestResearchColumn:
    """v1.4 新增：'我的调研结论'列（自由文本，第 9 列 I 列）。"""

    def test_调研列默认空_数据行(self, sample_results, tmp_path: Path):
        """item dict 无调研字段 → Excel 单元格默认空。"""
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=9)
            assert cell.value in (None, ""), \
                f"行{row_idx}: 调研列应为空，实际为 {cell.value!r}"

    def test_调研列表头_我的调研结论(self, sample_results, tmp_path: Path):
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        assert ws.cell(row=1, column=9).value == "我的调研结论"

    def test_调研列_行级着色_与同行建议列同色(self, sample_results, tmp_path: Path):
        """row_fill 覆盖到第 9 列：浅绿/浅黄/浅灰各跑一行。"""
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        # 建议行（第 4 行）→ 调研列应浅绿
        assert ws.cell(row=4, column=9).fill.fgColor.rgb in ("00E2EFDA", "E2EFDA", "FFE2EFDA")
        # 待研究行（第 2/5 行）→ 调研列应浅黄
        assert ws.cell(row=2, column=9).fill.fgColor.rgb in ("00FFF2CC", "FFF2CC", "FFFFF2CC")
        # 跳过行（第 3 行）→ 调研列应浅灰
        assert ws.cell(row=3, column=9).fill.fgColor.rgb in ("00F2F2F2", "F2F2F2", "FFF2F2F2")

    def test_调研列_自适应列宽(self, sample_results, tmp_path: Path):
        """不固定宽度，按内容自适应（用户手填文本长度不定）。"""
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        # 因为数据行默认空，列宽至少能容纳表头"我的调研结论"（6 个中文 = 12）
        # 调研列在第 9 列（I 列）
        col_width = ws.column_dimensions["I"].width
        assert col_width >= 12, f"调研列宽度至少 12（容纳表头），实际 {col_width}"

    def test_调研列_无DataValidation(self, sample_results, tmp_path: Path):
        """调研列是自由文本，不应加下拉框。"""
        out = tmp_path / "out.xlsx"
        generate(sample_results, out)
        ws = load_workbook(out).active
        dvs = list(ws.data_validations.dataValidation)
        # 唯一一条 DataValidation 应作用于 H 列（收藏），不应用于 I 列
        for dv in dvs:
            ranges = " ".join(str(s) for s in dv.sqref.ranges) if dv.sqref else ""
            # 只有"收藏"列（H 列）有 DataValidation
            assert "H2" in ranges
            # 不应有 I/J/K/L/M 列引用（其他列不应有 DataValidation）
            for col in ("I", "J", "K", "L", "M"):
                assert f"{col}2" not in ranges, \
                    f"调研列（I 列）不应有 DataValidation，但 {col}2 出现"

    def test_用户填写调研后_仍能正常生成(self, tmp_path: Path):
        """即使 item 没提供研究字段也不应崩 —— 用户只在 Excel 端手填。"""
        results = [{
            "keyword": "x",
            "translation": "x",
            "intent": "信息型",
            "type": "内容",
            "monetization": "AdSense",
            "advice": "建议",
            "target_user": "x",
            "urls": {
                "intitle": "http://x", "serp": "http://x",
                "trends": "http://x", "ahrefs": "http://x",
            },
        }]
        out = tmp_path / "out.xlsx"
        generate(results, out)  # 不应抛 KeyError
        ws = load_workbook(out).active
        # 调研列单元格保持空（实现不主动写入数据）
        assert ws.cell(row=2, column=9).value in (None, "")
