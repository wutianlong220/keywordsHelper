"""生成关键词分类 Excel。

实现 PRD 第 4.2 节 + Q3 调整 + v1.1 双轴升级 + v1.3 收藏列固定色 + v1.4 新增调研结果列 + v1.5 跳过（品牌）：
- 13 列固定顺序
- 行级着色（按「建议」：建议/待研究/跳过/跳过（品牌），覆盖 1~9 列；含新增"我的调研结论"列）
- 列级着色（链接列）
- 冻结首行 + 第 1、2 列
- 首行加筛选器
- 4 列可点击 URL（hyperlink）
- "收藏"列固定深绿背景 #548235 + 白字加粗
  - DataValidation 下拉框仅"收藏"一个选项
  - toggle 只决定"收藏"字是否出现，不影响背景色
- v1.4 新增"我的调研结论"列（自由文本，默认空，调研后手填备注，插入在"收藏"列右边）
  - 收藏列保持原位（第 8 列），链接列右移至 10-13
- v1.5 新增 advice="跳过（品牌）" 配色（复用跳过浅灰 #F2F2F2）

输入 results 是 list[dict]，每条 dict 字段：
  keyword, translation, intent, type, monetization, advice, target_user, urls
其中：
  type     - 关键词天然适合的网站类型（工具/内容/带货/游戏/导航/空）
  advice   - 给用户的行动建议（建议/待研究/跳过）
  urls     - dict：{"intitle", "serp", "trends", "ahrefs"}
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# ── 列定义（PRD 4.2 节 + Q3 调整 + v1.1 双轴 + v1.2 收藏列 + v1.4 调研结果列）──
# 列 4 = 类型（关键词天然适合什么网站）
# 列 6 = 建议（建议/待研究/跳过，给用户的行动建议）
# 列 8 = 收藏（下拉框），列 9 = 我的调研结论（自由文本），列 10-13 = 4 个链接列
_COLUMNS = [
    "英文原文",         # 1
    "中文翻译",         # 2
    "搜索意图",         # 3
    "类型",             # 4  ★ v1.1 新增：工具/内容/带货/游戏/导航/空
    "变现路径",         # 5
    "建议",             # 6  ★ 原"分类"改名 + 重定义为 3 档
    "目标用户",         # 7
    "收藏",             # 8  ★ v1.2：原"保留"改名 + 加条件格式（v1.4 后保持第 8 列）
    "我的调研结论",     # 9  ★ v1.4 新增：自由文本，调研后手填备注
    "intitle 检查",     # 10
    "SERP 检查",        # 11
    "谷歌趋势",         # 12
    "Ahrefs KD",        # 13
]

# 行级填充（按「建议」）：覆盖 1~9 列（含新增"我的调研结论"列，不覆盖收藏/链接列）
_ADVICE_FILLS = {
    "建议": PatternFill("solid", fgColor="E2EFDA"),    # 浅绿
    "待研究": PatternFill("solid", fgColor="FFF2CC"),  # 浅黄
    "跳过": PatternFill("solid", fgColor="F2F2F2"),    # 浅灰
    "跳过（品牌）": PatternFill("solid", fgColor="F2F2F2"),  # 浅灰（与"跳过"一致；v1.5 新增）
}

# 链接列填充（列 10~13，v1.4 后右移 1 位）
_LINK_FILLS = {
    10: PatternFill("solid", fgColor="DDEBF7"),  # intitle 浅蓝
    11: PatternFill("solid", fgColor="E4D7F0"),  # SERP 浅紫
    12: PatternFill("solid", fgColor="FCE4D6"),  # 趋势 浅橙
    13: PatternFill("solid", fgColor="E2EFDA"),  # Ahrefs 浅绿
}

# 链接列表头填充
_LINK_HEADER_FILLS = {
    10: PatternFill("solid", fgColor="4472C4"),
    11: PatternFill("solid", fgColor="7030A0"),
    12: PatternFill("solid", fgColor="C65911"),
    13: PatternFill("solid", fgColor="548235"),
}

_DEFAULT_HEADER_FILL = PatternFill("solid", fgColor="404040")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

_URL_KEYS = {10: "intitle", 11: "serp", 12: "trends", 13: "ahrefs"}
_HYPERLINK_FONT = Font(color="0563C1", underline="single")

# 列宽常量
_URL_WIDTH = 12        # 4 个链接列固定 12 字符
_FAVORITE_WIDTH = 8     # "收藏"列固定 8 字符
_FAVORITE_COL = 8        # "收藏"列位于第 8 列（v1.4 新增调研列插到右边后，收藏不挪位）

# 收藏列固定深绿背景 + 白字加粗
# v1.3 决策：toggle 只决定"收藏"文本是否出现，不再切背景色
_FAVORITE_BG_FILL = PatternFill("solid", fgColor="548235")
_FAVORITE_FONT = Font(bold=True, color="FFFFFF")
_KEYWORD_WIDTH = 30     # "英文原文"列固定 30 字符（与中文翻译列对齐）


def generate(results: list[dict], output_path: str | Path) -> Path:
    """生成 Excel 文件，返回输出路径 Path 对象。

    自动创建输出目录（如不存在）。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "关键词分类"

    _write_header(ws)
    _write_rows(ws, results)
    _apply_layout(ws)
    # 仅当有数据行时才加 DataValidation（空 results 时列号非法）
    # v1.3：不再用条件格式，背景色由 _write_rows 直接涂
    if results:
        _add_data_validation(ws, last_row=len(results) + 1)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _write_header(ws):
    """写表头 13 列 + 着色。"""
    for col_idx, name in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = _LINK_HEADER_FILLS.get(col_idx, _DEFAULT_HEADER_FILL)
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN


def _write_rows(ws, results: list[dict]):
    """写数据行。"""
    for row_idx, item in enumerate(results, 2):
        advice = item["advice"]
        row_fill = _ADVICE_FILLS.get(advice)

        # 1~9 列：行级填充（含 v1.4 新增"我的调研结论"列，不含收藏列 8 —— 因收藏列有固定深绿）
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_idx, value=_cell_value(item, col_idx))
            if row_fill:
                cell.fill = row_fill

        # 列 8 = 收藏（固定深绿背景 + 白字加粗；toggle 仅决定文本是否出现）
        fav_cell = ws.cell(row=row_idx, column=_FAVORITE_COL, value="")
        fav_cell.fill = _FAVORITE_BG_FILL
        fav_cell.font = _FAVORITE_FONT

        # 10~13 列：URL（覆盖行级填充）
        for col_idx, url_key in _URL_KEYS.items():
            url = item["urls"][url_key]
            cell = ws.cell(row=row_idx, column=col_idx, value=url)
            cell.hyperlink = url
            cell.font = _HYPERLINK_FONT
            cell.fill = _LINK_FILLS[col_idx]


def _cell_value(item: dict, col_idx: int):
    """根据列号从 item 中取值（col_idx 范围 1~9）。

    特殊处理：
    - 「类型」列里的"空"是哨兵值，转为空字符串让 Excel 单元格留空。
    - 「收藏」「我的调研结论」列无对应字段，默认返回空串。
    """
    keys = [
        "keyword",       # 1
        "translation",   # 2
        "intent",        # 3
        "type",          # 4 ★ v1.1 新增
        "monetization",  # 5
        "advice",        # 6 ★ 原"category"改名
        "target_user",   # 7
        None,            # 8 = 收藏（单独写，不读 item）
        None,            # 9 = 我的调研结论（用户手填，不读 item）★ v1.4
    ]
    key = keys[col_idx - 1]
    if key is None:
        return ""
    val = item.get(key, "")
    # 类型列：JSON 里 "空" = 无法归类；Excel 上表现为空白单元格
    if key == "type" and val == "空":
        return ""
    return val


def _add_data_validation(ws, last_row: int):
    """给"收藏"列加下拉框（v1.5：含"收藏 / 丢弃"两个选项，均手动独立）。

    行为：
    - 默认空 = 不标记
    - 下拉选"收藏" = 标记为收藏
    - 下拉选"丢弃" = 标记为丢弃
    - 单元格内容清空（Delete 键）= 取消标记
    - v1.5：与「建议」列不联动，由用户手动区分
    """
    dv = DataValidation(
        type="list",
        formula1='"收藏,丢弃"',  # v1.5：新增"丢弃"选项
        allow_blank=True,
        showErrorMessage=True,
    )
    dv.prompt = "选择『收藏』收藏此词，选择『丢弃』丢弃此词，留空表示不标记"
    dv.promptTitle = "标记词"
    dv.error = "请选择『收藏』『丢弃』或留空"
    dv.errorTitle = "无效输入"
    ws.add_data_validation(dv)
    # v1.4：收藏列仍在第 8 列（H 列），不变
    col_letter = get_column_letter(_FAVORITE_COL)
    dv.add(f"{col_letter}2:{col_letter}{last_row}")


def _apply_layout(ws):
    """冻结 + 筛选器 + 列宽。"""
    # 冻结首行 + 第 1、2 列
    ws.freeze_panes = "C2"

    # 筛选器（首行）
    last_col = get_column_letter(len(_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}1"

    # 列宽处理
    for col_idx in range(1, len(_COLUMNS) + 1):
        # 列 1 = "英文原文"：固定 30 字符（与中文翻译列对齐，避免英文长词撑爆列宽）
        if col_idx == 1:
            ws.column_dimensions[get_column_letter(col_idx)].width = _KEYWORD_WIDTH
            continue

        # 列 _FAVORITE_COL = "收藏"：固定 8 字符
        if col_idx == _FAVORITE_COL:
            ws.column_dimensions[get_column_letter(col_idx)].width = _FAVORITE_WIDTH
            continue

        # 链接列：固定 12 字符
        if col_idx in _URL_KEYS:
            ws.column_dimensions[get_column_letter(col_idx)].width = _URL_WIDTH
            continue

        # 其他列自适应（含"我的调研结论"列：按手填内容长度自适应）
        header_len = _approx_width(_COLUMNS[col_idx - 1])
        max_len = header_len
        for row_idx in range(2, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            length = _approx_width(str(v))
            if length > max_len:
                max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


def _approx_width(s: str) -> int:
    """粗略估算字符串在 Excel 中的显示宽度（中文字符按 2 计）。"""
    width = 0
    for ch in s:
        if ord(ch) > 127:
            width += 2
        else:
            width += 1
    return width
